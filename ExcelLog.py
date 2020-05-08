import openpyxl

#set path, open file, specify the sheet
path = "C://Users//Mikhail//PycharmProjects//SeleniumTest//Book1.xlsx"
book = openpyxl.load_workbook(path)
sheet = book.active #book.get_sheet_by_name("List1") if more than one

rows = sheet.max_row #5
columns = sheet.max_column #4

#print(rows)
#print(columns)

for r in range (1, rows + 1):
    for c in range (1, columns + 1):
        print(sheet.cell(row=r, column=c).value, end= "     ")
    print()

#result
#ID     FirstName     LastName     Salary
#1     Bono     Voice     100000
#2     Edge     Guitar     200000
#3     Adam     Bass     300000
#4     Larry     Drum     400000